"""
Metals & Mining data engine — Yalchin Nijat
Pulls prices + financial statements for First Majestic and peers,
recomputes the MGT 231 ratio framework, writes a comps workbook.

To refresh everything:  python metals_engine.py
Output: comps_workbook.xlsx (in this folder)
"""
import datetime
import json
import pathlib
import pandas as pd
import yfinance as yf
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

CALL_DATE = "2024-11-21"          # date of the original MGT 231 analysis
TODAY = datetime.date.today().isoformat()

# ticker: (display name, statement currency)
COMPANIES = {
    "AG":    ("First Majestic Silver", "USD"),
    "PAAS":  ("Pan American Silver",   "USD"),
    "FNLPF": ("Fresnillo PLC",         "GBP"),
    "HL":    ("Hecla Mining",          "USD"),
    "CDE":   ("Coeur Mining",          "USD"),
    "EXK":   ("Endeavour Silver",      "USD"),
    "FSM":   ("Fortuna Mining",        "USD"),
    "SVM":   ("Silvercorp Metals",     "USD"),
    "AYA.TO":("Aya Gold & Silver",     "CAD"),
    "ASM":   ("Avino Silver & Gold",   "USD"),
}
SILVER = "SI=F"   # front-month silver futures

BLUE = Font(color="0000FF")            # hardcoded inputs
GREEN = Font(color="008000")           # cross-sheet formula links
BOLD = Font(bold=True)
HDR_FILL = PatternFill("solid", start_color="D9E1F2")
YELLOW = PatternFill("solid", start_color="FFFF00")

def pick(df, labels):
    """Return the first matching row label's series from a statement, else None."""
    for lab in labels:
        if lab in df.index:
            return df.loc[lab]
    return None

def val(series, col):
    if series is None or col not in series.index:
        return None
    v = series[col]
    return None if pd.isna(v) else float(v) / 1e6   # millions

def fetch_company(tkr):
    t = yf.Ticker(tkr)
    inc, bs, cf = t.financials, t.balance_sheet, t.cashflow
    rows = []
    for col in inc.columns:
        year = col.year
        if val(pick(inc, ["Total Revenue"]), col) is None:
            continue   # yfinance pads the oldest year with an all-empty column
        rows.append({
            "Ticker": tkr, "Year": year,
            "Revenue":  val(pick(inc, ["Total Revenue"]), col),
            "EBIT":     val(pick(inc, ["EBIT", "Operating Income"]), col),
            "EBITDA":   val(pick(inc, ["EBITDA", "Normalized EBITDA"]), col),
            "NetIncome":val(pick(inc, ["Net Income", "Net Income Common Stockholders"]), col),
            "IntExp":   val(pick(inc, ["Interest Expense"]), col),
            "TotDebt":  val(pick(bs, ["Total Debt"]), col) if col in bs.columns else None,
            "Equity":   val(pick(bs, ["Stockholders Equity", "Common Stock Equity"]), col) if col in bs.columns else None,
            "TotAssets":val(pick(bs, ["Total Assets"]), col) if col in bs.columns else None,
            "CurAssets":val(pick(bs, ["Current Assets"]), col) if col in bs.columns else None,
            "Inventory":val(pick(bs, ["Inventory"]), col) if col in bs.columns else None,
            "CurLiab":  val(pick(bs, ["Current Liabilities"]), col) if col in bs.columns else None,
            "FCF":      val(pick(cf, ["Free Cash Flow"]), col) if col in cf.columns else None,
        })
    info = t.info or {}
    snap = {
        "Ticker": tkr,
        "Price": info.get("currentPrice") or info.get("regularMarketPrice"),
        "MktCap_M": (info.get("marketCap") or 0) / 1e6 or None,
        "BookValuePS": info.get("bookValue"),
        "EV_M": (info.get("enterpriseValue") or 0) / 1e6 or None,
        "TrailEBITDA_M": (info.get("ebitda") or 0) / 1e6 or None,
        "Shares_M": (info.get("sharesOutstanding") or 0) / 1e6 or None,
    }
    return rows, snap

def fetch_returns():
    out = []
    for tkr, (name, ccy) in list(COMPANIES.items()) + [(SILVER, ("Silver (spot proxy)", "USD"))]:
        h = yf.Ticker(tkr).history(start=CALL_DATE)
        if len(h) == 0:
            continue
        out.append({"Ticker": tkr, "Name": name,
                    "PriceAtCall": round(float(h.Close.iloc[0]), 2),
                    "PriceNow": round(float(h.Close.iloc[-1]), 2),
                    "AsOf": str(h.index[-1].date())})
    return out

def main():
    print("Fetching returns since", CALL_DATE, "...")
    returns = fetch_returns()
    print("Fetching statements for", len(COMPANIES), "companies ...")
    raw, snaps = [], []
    for tkr in COMPANIES:
        try:
            rows, snap = fetch_company(tkr)
            raw += rows
            snaps.append(snap)
            print("  ok:", tkr)
        except Exception as e:
            print("  FAILED:", tkr, "-", e)

    wb = Workbook()

    # ---- README ----
    ws = wb.active
    ws.title = "README"
    ws["A1"] = "Metals & Mining Comps Workbook"
    ws["A1"].font = Font(bold=True, size=14)
    lines = [
        f"Generated {TODAY} by metals_engine.py. To refresh: open a terminal in this folder and run:  python metals_engine.py",
        "Sheets: Returns (performance since the Nov 21 2024 call) | Raw Data (statement inputs, blue = hardcoded from Yahoo Finance)",
        "Ratios (the MGT 231 framework, green = formulas linking to Raw Data) | Valuation (current multiples) | AISC (manual, from company reports)",
        "Ratios are within-company so statement currency (GBP for Fresnillo, CAD for Aya) does not distort comparisons.",
        "Data source: Yahoo Finance via the yfinance library. Verify key figures against company filings before citing.",
    ]
    for i, ln in enumerate(lines, start=3):
        ws.cell(row=i, column=1, value=ln)
    ws.column_dimensions["A"].width = 120

    # ---- Returns ----
    ws = wb.create_sheet("Returns")
    hdr = ["Name", "Ticker", f"Price {CALL_DATE}", "Price now", "As of", "Total return"]
    for c, h in enumerate(hdr, 1):
        cell = ws.cell(row=1, column=c, value=h); cell.font = BOLD; cell.fill = HDR_FILL
    for r, row in enumerate(returns, 2):
        ws.cell(row=r, column=1, value=row["Name"])
        ws.cell(row=r, column=2, value=row["Ticker"])
        ws.cell(row=r, column=3, value=row["PriceAtCall"]).font = BLUE
        ws.cell(row=r, column=4, value=row["PriceNow"]).font = BLUE
        ws.cell(row=r, column=5, value=row["AsOf"])
        f = ws.cell(row=r, column=6, value=f"=IFERROR(D{r}/C{r}-1,\"n/m\")")
        f.number_format = "0.0%"
    for c, w in zip("ABCDEF", [26, 10, 16, 12, 12, 12]):
        ws.column_dimensions[c].width = w

    # ---- Raw Data (long format: one row per company-year) ----
    ws = wb.create_sheet("Raw Data")
    cols = ["Ticker", "Year", "Revenue", "EBIT", "EBITDA", "NetIncome", "IntExp",
            "TotDebt", "Equity", "TotAssets", "CurAssets", "Inventory", "CurLiab", "FCF"]
    ws.cell(row=1, column=1, value="All statement figures in millions of the company's reporting currency (source: Yahoo Finance)").font = Font(italic=True)
    for c, h in enumerate(cols, 1):
        cell = ws.cell(row=2, column=c, value=h); cell.font = BOLD; cell.fill = HDR_FILL
    raw_sorted = sorted(raw, key=lambda r: (r["Ticker"], -r["Year"]))
    for r, row in enumerate(raw_sorted, 3):
        for c, k in enumerate(cols, 1):
            cell = ws.cell(row=r, column=c, value=row[k])
            if k not in ("Ticker", "Year"):
                cell.font = BLUE
                cell.number_format = "#,##0;(#,##0);-"
            if k == "Year":
                cell.number_format = "0"
    for i in range(1, len(cols) + 1):
        ws.column_dimensions[get_column_letter(i)].width = 11

    # ---- Ratios (formulas referencing Raw Data, same row order) ----
    ws = wb.create_sheet("Ratios")
    rcols = ["Ticker", "Year", "EBIT Margin", "FCF Margin", "Quick Ratio",
             "Debt/Capital", "Debt/EBITDA", "Interest Coverage",
             "Net Margin", "Asset Turnover", "Equity Multiplier", "ROE (DuPont)"]
    for c, h in enumerate(rcols, 1):
        cell = ws.cell(row=1, column=c, value=h); cell.font = BOLD; cell.fill = HDR_FILL
    n = len(raw_sorted)
    # Raw Data data rows start at 3; Ratios data rows start at 2 → offset +1
    for i in range(n):
        rr, tr = i + 3, i + 2
        D = f"'Raw Data'!"
        ws.cell(row=tr, column=1, value=f"={D}A{rr}")
        ws.cell(row=tr, column=2, value=f"={D}B{rr}").number_format = "0"
        def guarded(cells, expr):
            # show n/a when any input cell is blank instead of silently computing with zeros
            refs = ",".join(f"{D}{c}{rr}" for c in cells)
            return f"=IF(COUNT({refs})<{len(cells)},\"n/a\",{expr})"
        formulas = [
            (3,  guarded("DC", f"{D}D{rr}/{D}C{rr}"), "0.0%"),                       # EBIT margin
            (4,  guarded("NC", f"{D}N{rr}/{D}C{rr}"), "0.0%"),                       # FCF margin
            (5,  guarded("KLM", f"({D}K{rr}-{D}L{rr})/{D}M{rr}"), "0.00"),           # quick
            (6,  guarded("HI", f"{D}H{rr}/({D}H{rr}+{D}I{rr})"), "0.0%"),            # debt/cap
            (7,  guarded("HE", f"IF({D}E{rr}<=0,\"n/m\",{D}H{rr}/{D}E{rr})"), "0.0x"),          # debt/EBITDA
            (8,  guarded("DG", f"IF({D}G{rr}=0,\"n/m\",{D}D{rr}/{D}G{rr})"), "0.0x"),           # int cov
            (9,  guarded("FC", f"{D}F{rr}/{D}C{rr}"), "0.0%"),                       # net margin
            (10, guarded("CJ", f"{D}C{rr}/{D}J{rr}"), "0.00"),                       # asset turnover
            (11, guarded("JI", f"{D}J{rr}/{D}I{rr}"), "0.00"),                       # equity multiplier
            (12, f"=IF(COUNT(I{tr},J{tr},K{tr})<3,\"n/a\",I{tr}*J{tr}*K{tr})", "0.0%"),  # ROE = product
        ]
        for col, f, fmt in formulas:
            cell = ws.cell(row=tr, column=col, value=f)
            cell.number_format = fmt
            cell.font = GREEN if col < 12 else Font(color="000000")
    for i in range(1, len(rcols) + 1):
        ws.column_dimensions[get_column_letter(i)].width = 13

    # ---- Valuation snapshot ----
    ws = wb.create_sheet("Valuation")
    vcols = ["Name", "Ticker", "Price", "Mkt Cap ($mm)", "Book Value/sh",
             "EV ($mm)", "Trailing EBITDA ($mm)", "P/B", "EV/EBITDA"]
    ws.cell(row=1, column=1, value=f"Snapshot {TODAY} — listed-currency figures from Yahoo Finance").font = Font(italic=True)
    for c, h in enumerate(vcols, 1):
        cell = ws.cell(row=2, column=c, value=h); cell.font = BOLD; cell.fill = HDR_FILL
    for r, s in enumerate(snaps, 3):
        name = COMPANIES[s["Ticker"]][0]
        ws.cell(row=r, column=1, value=name)
        ws.cell(row=r, column=2, value=s["Ticker"])
        for c, k, fmt in [(3, "Price", "#,##0.00"), (4, "MktCap_M", "#,##0"),
                          (5, "BookValuePS", "#,##0.00"), (6, "EV_M", "#,##0"),
                          (7, "TrailEBITDA_M", "#,##0")]:
            cell = ws.cell(row=r, column=c, value=s[k]); cell.font = BLUE; cell.number_format = fmt
        ws.cell(row=r, column=8, value=f"=IFERROR(C{r}/E{r},\"n/m\")").number_format = "0.0x"
        ws.cell(row=r, column=9, value=f"=IF(G{r}<=0,\"n/m\",F{r}/G{r})").number_format = "0.0x"
    for i, w in enumerate([26, 10, 10, 13, 13, 12, 18, 8, 10], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ---- AISC (manual inputs from company reports — no free API has these) ----
    ws = wb.create_sheet("AISC")
    ws["A1"] = "AISC per silver-equivalent ounce (US$) — manual inputs from company press releases; yellow = fill in"
    ws["A1"].font = Font(italic=True)
    for c, h in enumerate(["Company", "FY2025 AISC $/AgEq oz", "Source"], 1):
        cell = ws.cell(row=2, column=c, value=h); cell.font = BOLD; cell.fill = HDR_FILL
    ws["A3"] = "First Majestic Silver"
    ws["B3"] = 21.17; ws["B3"].font = BLUE
    ws["C3"] = "FY2025 results press release, firstmajestic.com"
    row_i = 4
    for tkr, (name, _) in COMPANIES.items():
        if tkr == "AG":
            continue
        ws.cell(row=row_i, column=1, value=name)
        ws.cell(row=row_i, column=2).fill = YELLOW
        row_i += 1
    for c, w in zip("ABC", [26, 22, 50]):
        ws.column_dimensions[c].width = w

    out = "comps_workbook.xlsx"
    wb.save(out)
    print("Saved", out)
    write_site_data(returns, raw_sorted, snaps)

AISC_MANUAL = {  # US$/AgEq oz, from company press releases — update by hand
    "AG": {"2025": 21.17, "source": "FY2025 results, firstmajestic.com"},
}

def ratio_block(d):
    """Compute display ratios from one company-year of raw data (None-safe)."""
    def div(a, b):
        return None if (a is None or not b) else a / b
    quick = None
    if all(d.get(k) is not None for k in ("CurAssets", "Inventory", "CurLiab")) and d["CurLiab"]:
        quick = (d["CurAssets"] - d["Inventory"]) / d["CurLiab"]
    de = None
    if d.get("TotDebt") is not None and d.get("EBITDA") and d["EBITDA"] > 0:
        de = d["TotDebt"] / d["EBITDA"]
    return {
        "ebitMargin": div(d.get("EBIT"), d.get("Revenue")),
        "fcfMargin": div(d.get("FCF"), d.get("Revenue")),
        "netMargin": div(d.get("NetIncome"), d.get("Revenue")),
        "roe": div(d.get("NetIncome"), d.get("Equity")),
        "quick": quick,
        "debtEbitda": de,
        "debtCap": div(d.get("TotDebt"), (d.get("TotDebt") or 0) + (d.get("Equity") or 0) or None),
    }

def write_site_data(returns, raw_sorted, snaps):
    """Emit docs/data.js so the website shows the same numbers as the workbook."""
    latest, ag_series = {}, []
    for row in raw_sorted:
        t, y = row["Ticker"], row["Year"]
        if t not in latest or y > latest[t]["year"]:
            latest[t] = {"year": y, **ratio_block(row)}
        if t == "AG":
            ag_series.append({"year": y, **ratio_block(row)})
    data = {
        "generated": TODAY,
        "callDate": CALL_DATE,
        "companies": {t: {"name": n, "ccy": c} for t, (n, c) in COMPANIES.items()},
        "returns": returns,
        "latestRatios": latest,
        "agHistory": sorted(ag_series, key=lambda r: r["year"]),
        "valuation": snaps,
        "aisc": AISC_MANUAL,
    }
    site = pathlib.Path(__file__).resolve().parent.parent / "docs"
    site.mkdir(exist_ok=True)
    (site / "data.js").write_text(
        "window.SITE_DATA = " + json.dumps(data, indent=1) + ";\n", encoding="utf-8")
    print("Saved", site / "data.js")

if __name__ == "__main__":
    main()
