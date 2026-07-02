"""
Builds FirstMajestic_NAV.xlsx — a mine-by-mine net-asset-value model.
Regenerate after changing this script:  python build_nav_model.py
All numbers dated July 2, 2026. Yellow cells in the workbook = assumptions to revisit.
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

BLUE = Font(color="0000FF"); BOLD = Font(bold=True); GREEN = Font(color="008000")
YEL = PatternFill("solid", start_color="FFFF00"); HDR = PatternFill("solid", start_color="D9E1F2")
YEARS = list(range(2026, 2036))            # 10-year horizon, columns D..M
COL0 = 4                                    # first year column = D

# ---- inputs gathered July 2, 2026 (sources in README sheet) ----
DECKS_AG = {1: [55,50,47,45,45,45,45,45,45,45],      # conservative / bank-fade (HSBC-style)
            2: [61]*10,                               # spot-anchored flat
            3: [62,72,80,75,70,70,70,70,70,70]}       # structural-deficit bull
DECKS_AU = {1: [4000,3800,3600,3500,3500,3500,3500,3500,3500,3500],
            2: [4120]*10,
            3: [4200,4500,4800,4500,4400,4400,4400,4400,4400,4400]}
# name, Ag Moz/yr, Au koz/yr, byproduct AgEq Moz/yr (Zn/Pb), AISC $/AgEq, attrib P&P AgEq Moz, life-extension yrs
MINES = [
    ("San Dimas",      4.2, 52.0, 0.0, 25.5, 22.0, 3),
    ("Santa Elena",    1.4, 67.5, 0.0, 25.7, 48.4, 2),
    ("Los Gatos (70%)",5.1,  3.0, 3.0, 18.5, 55.3, 3),
    ("La Encantada",   2.95, 0.0, 0.2, 31.5, 13.5, 1),
]
JC = dict(au_koz=120, aisc_au=1900, start=2028, years=8, capex26=40, capex27=35)
CORP = dict(disc=0.05, tax_mx=0.40, tax_us=0.25, shares=488.7, netcash=687.4,
            ga=60, ga_years=10, price=17.42, jc_on=1, deck=2)

wb = Workbook()

# ================= README =================
ws = wb.active; ws.title = "README"
ws["A1"] = "First Majestic Silver — Net Asset Value model"; ws["A1"].font = Font(bold=True, size=14)
notes = [
 "WHAT THIS IS: each producing mine is a mini-DCF — project production, price it with the silver/gold deck,",
 "subtract all-in sustaining costs and tax, discount each year's cash flow, sum. Add net cash, subtract",
 "corporate overhead, divide by shares = NAV per share. Compare to the market price = P/NAV.",
 "",
 "HOW TO USE: everything you might change lives on the Assumptions sheet. Yellow = judgment calls to own.",
 "Blue = hardcoded facts from disclosures. Black/green = formulas. Change the deck number or discount rate",
 "and the whole model reprices instantly.",
 "",
 "KEY SIMPLIFICATIONS (know these for interviews):",
 "1. Flat production over each mine's modeled life (real mines have grade/throughput profiles).",
 "2. AISC used as the all-in cost proxy — it includes sustaining capex but NOT expansionary capex",
 "   (company guides $154-171M expansionary in 2026 alone). NAV is optimistic by roughly that PV.",
 "3. Mine life = reserve life + a judgment 'resource conversion' extension. Defend the extension years.",
 "4. Los Gatos zinc/lead byproducts modeled as an AgEq stream priced at silver.",
 "5. Jerritt Canyon modeled as a simple restart case (guides say H2 2027 production, PFS due Q4 2026 —",
 "   update this block when the study is out). Toggle it on/off on Assumptions.",
 "6. First Mint, exploration properties, and NCI subtleties beyond the 70% Gatos share: excluded.",
 "",
 "MID-YEAR CONVENTION: cash flows discounted at (t - 0.5) — cash arrives through the year, not Dec 31.",
 "P/NAV INTUITION: quality silver producers historically trade 0.8x-1.5x NAV(5%); the premium/discount",
 "reflects mine-life optionality, jurisdiction, and balance-sheet quality.",
 "",
 "DATA SOURCES: 2026 guidance & AISC by mine (Jan 2026 outlook release); P&P reserves Dec 31 2025",
 "(Mar 2026 release); Q1 2026 balance sheet (cash $984.8M, debt $297.4M, shares 488.7M);",
 "spot Ag ~$61 / Au ~$4,120 (Jul 2, 2026); decks bracket HSBC / spot / JPM-GS style forecasts.",
]
for i, t in enumerate(notes, 3): ws.cell(row=i, column=1, value=t)
ws.column_dimensions["A"].width = 110

# ================= Assumptions =================
ws = wb.create_sheet("Assumptions")
ws["A1"] = "Assumptions — yellow cells are yours to defend"; ws["A1"].font = Font(bold=True, size=13)

ws["A3"] = "Selected price deck (1=Conservative, 2=Spot-flat, 3=Bull)"; ws["B3"] = CORP["deck"]
ws["B3"].fill = YEL; ws["B3"].font = BLUE

ws["A5"] = "Price decks (US$/oz)"; ws["A5"].font = BOLD
for j, y in enumerate(YEARS):
    c = ws.cell(row=5, column=COL0+j, value=y); c.font = BOLD; c.fill = HDR; c.number_format = "0"
deck_rows_ag, deck_rows_au = {}, {}
labels = {1: "Deck 1 — Conservative (bank-fade)", 2: "Deck 2 — Spot-anchored flat", 3: "Deck 3 — Structural bull"}
r = 6
for d in (1, 2, 3):
    ws.cell(row=r, column=1, value=f"Silver: {labels[d]}")
    for j, v in enumerate(DECKS_AG[d]):
        c = ws.cell(row=r, column=COL0+j, value=v); c.font = BLUE; c.fill = YEL; c.number_format = "#,##0"
    deck_rows_ag[d] = r; r += 1
ws.cell(row=r, column=1, value="Silver price USED (per selected deck)").font = BOLD
for j in range(10):
    L = get_column_letter(COL0+j)
    ws.cell(row=r, column=COL0+j, value=f"=CHOOSE($B$3,{L}6,{L}7,{L}8)").number_format = "#,##0"
AG_ROW = r; r += 2
for d in (1, 2, 3):
    ws.cell(row=r, column=1, value=f"Gold: {labels[d]}")
    for j, v in enumerate(DECKS_AU[d]):
        c = ws.cell(row=r, column=COL0+j, value=v); c.font = BLUE; c.fill = YEL; c.number_format = "#,##0"
    deck_rows_au[d] = r; r += 1
ws.cell(row=r, column=1, value="Gold price USED (per selected deck)").font = BOLD
for j in range(10):
    L = get_column_letter(COL0+j)
    ws.cell(row=r, column=COL0+j, value=f"=CHOOSE($B$3,{L}{r-3},{L}{r-2},{L}{r-1})").number_format = "#,##0"
AU_ROW = r

r += 2
corp_items = [("Discount rate (NAV convention: 5% for precious)", CORP["disc"], "0.0%", True),
              ("Tax rate — Mexico (income + mining duties, effective)", CORP["tax_mx"], "0.0%", True),
              ("Tax rate — US / Nevada (effective)", CORP["tax_us"], "0.0%", True),
              ("Shares outstanding (M)", CORP["shares"], "#,##0.0", False),
              ("Net cash, Q1 2026 ($mm) = 984.8 - 297.4", CORP["netcash"], "#,##0", False),
              ("Corporate G&A ($mm/yr)", CORP["ga"], "#,##0", True),
              ("G&A years capitalized", CORP["ga_years"], "0", True),
              ("Current share price (US$, Jul 2 2026)", CORP["price"], "#,##0.00", False),
              ("Include Jerritt Canyon? (1=yes, 0=no)", CORP["jc_on"], "0", True)]
CORP_ROW0 = r
for i, (lab, v, fmt, yellow) in enumerate(corp_items):
    ws.cell(row=r+i, column=1, value=lab)
    c = ws.cell(row=r+i, column=2, value=v); c.font = BLUE; c.number_format = fmt
    if yellow: c.fill = YEL
R_DISC, R_TAXMX, R_TAXUS, R_SHARES, R_NETCASH, R_GA, R_GAYRS, R_PRICE, R_JCON = [CORP_ROW0+i for i in range(9)]

r = CORP_ROW0 + 11
ws.cell(row=r, column=1, value="Mine parameters").font = BOLD
hdrs = ["Mine", "Ag Moz/yr", "Au koz/yr", "Byprod AgEq Moz/yr", "AISC $/AgEq oz",
        "P&P reserves attrib (AgEq Moz)", "Reserve life (yrs)", "+ Conversion yrs (judgment)", "Modeled life (yrs)"]
for c_i, h in enumerate(hdrs, 1):
    c = ws.cell(row=r+1, column=c_i, value=h); c.font = BOLD; c.fill = HDR
    c.alignment = Alignment(wrap_text=True)
MINE_ROW0 = r + 2
for i, (nm, ag, au, byp, aisc, rsv, ext) in enumerate(MINES):
    rr = MINE_ROW0 + i
    ws.cell(row=rr, column=1, value=nm)
    for c_i, v, fmt in [(2, ag, "0.00"), (3, au, "0.0"), (4, byp, "0.00"), (5, aisc, "0.00"), (6, rsv, "0.0")]:
        c = ws.cell(row=rr, column=c_i, value=v); c.font = BLUE; c.number_format = fmt
    ws.cell(row=rr, column=7, value=f"=F{rr}/(B{rr}+C{rr}/1000*67.5+D{rr})").number_format = "0.0"
    c = ws.cell(row=rr, column=8, value=ext); c.font = BLUE; c.fill = YEL; c.number_format = "0"
    ws.cell(row=rr, column=9, value=f"=ROUND(G{rr},0)+H{rr}").number_format = "0"

r = MINE_ROW0 + len(MINES) + 1
ws.cell(row=r, column=1, value="Jerritt Canyon restart case (update after Q4 2026 pre-feasibility study)").font = BOLD
jc_items = [("Gold production (koz/yr)", JC["au_koz"], "0"), ("AISC ($/Au oz)", JC["aisc_au"], "#,##0"),
            ("First production year", JC["start"], "0"), ("Production years modeled", JC["years"], "0"),
            ("Restart capex 2026 ($mm)", JC["capex26"], "#,##0"), ("Restart capex 2027 ($mm)", JC["capex27"], "#,##0")]
JC_ROW0 = r + 1
for i, (lab, v, fmt) in enumerate(jc_items):
    ws.cell(row=JC_ROW0+i, column=1, value=lab)
    c = ws.cell(row=JC_ROW0+i, column=2, value=v); c.font = BLUE; c.fill = YEL; c.number_format = fmt
ws.column_dimensions["A"].width = 48
for c_i in range(2, 14): ws.column_dimensions[get_column_letter(c_i)].width = 13

A = "Assumptions!"
# ================= mine sheets =================
def mine_sheet(name, idx):
    p = MINE_ROW0 + idx          # parameter row on Assumptions
    s = wb.create_sheet(name[:28].replace("(", "").replace(")", "").replace("%", ""))
    s["A1"] = f"{name} — mini-DCF"; s["A1"].font = Font(bold=True, size=12)
    rows = ["Year", "t", "Active (1/0)", "Ag prod (Moz)", "Au prod (koz)", "Byprod AgEq (Moz)",
            "Ag price", "Au price", "AgEq total (Moz)", "Revenue ($mm)", "AISC cost ($mm)",
            "Pre-tax CF ($mm)", "Tax ($mm)", "After-tax CF ($mm)", "Discount factor", "PV ($mm)"]
    for i, lab in enumerate(rows, 3): s.cell(row=i, column=1, value=lab)
    for j, y in enumerate(YEARS):
        cL = get_column_letter(COL0+j)
        s.cell(row=3, column=COL0+j, value=y).font = BOLD
        s.cell(row=4, column=COL0+j, value=j+1)
        s.cell(row=5, column=COL0+j, value=f"=IF({cL}4<={A}$I${p},1,0)")
        s.cell(row=6, column=COL0+j, value=f"={A}$B${p}*{cL}5").number_format = "0.00"
        s.cell(row=7, column=COL0+j, value=f"={A}$C${p}*{cL}5").number_format = "0.0"
        s.cell(row=8, column=COL0+j, value=f"={A}$D${p}*{cL}5").number_format = "0.00"
        s.cell(row=9, column=COL0+j, value=f"={A}{cL}${AG_ROW}").number_format = "#,##0"
        s.cell(row=10, column=COL0+j, value=f"={A}{cL}${AU_ROW}").number_format = "#,##0"
        s.cell(row=11, column=COL0+j, value=f"={cL}6+{cL}7/1000*{cL}10/{cL}9+{cL}8").number_format = "0.00"
        s.cell(row=12, column=COL0+j, value=f"={cL}6*{cL}9+{cL}7/1000*{cL}10+{cL}8*{cL}9").number_format = "#,##0"
        s.cell(row=13, column=COL0+j, value=f"={A}$E${p}*{cL}11").number_format = "#,##0"
        s.cell(row=14, column=COL0+j, value=f"={cL}12-{cL}13").number_format = "#,##0;(#,##0)"
        s.cell(row=15, column=COL0+j, value=f"=MAX(0,{cL}14*{A}$B${R_TAXMX})").number_format = "#,##0"
        s.cell(row=16, column=COL0+j, value=f"={cL}14-{cL}15").number_format = "#,##0;(#,##0)"
        s.cell(row=17, column=COL0+j, value=f"=1/(1+{A}$B${R_DISC})^({cL}4-0.5)").number_format = "0.000"
        s.cell(row=18, column=COL0+j, value=f"={cL}16*{cL}17").number_format = "#,##0;(#,##0)"
    s["A20"] = "NPV ($mm)"; s["A20"].font = BOLD
    s["B20"] = f"=SUM(D18:{get_column_letter(COL0+9)}18)"; s["B20"].font = BOLD; s["B20"].number_format = "#,##0"
    s.column_dimensions["A"].width = 22
    return s.title

mine_tabs = [mine_sheet(m[0], i) for i, m in enumerate(MINES)]

# Jerritt Canyon sheet
s = wb.create_sheet("JerrittCanyon")
s["A1"] = "Jerritt Canyon restart — simple case (revisit after Q4 2026 PFS)"; s["A1"].font = Font(bold=True, size=12)
labs = ["Year", "t", "Producing (1/0)", "Au prod (koz)", "Au price", "Margin $/oz (Au price - AISC)",
        "Operating CF ($mm)", "Restart capex ($mm)", "Pre-tax CF ($mm)", "Tax ($mm)",
        "After-tax CF ($mm)", "Discount factor", "PV ($mm)"]
for i, lab in enumerate(labs, 3): s.cell(row=i, column=1, value=lab)
for j, y in enumerate(YEARS):
    cL = get_column_letter(COL0+j)
    s.cell(row=3, column=COL0+j, value=y).font = BOLD
    s.cell(row=4, column=COL0+j, value=j+1)
    s.cell(row=5, column=COL0+j, value=f"=IF(AND({cL}3>={A}$B${JC_ROW0+2},{cL}3<{A}$B${JC_ROW0+2}+{A}$B${JC_ROW0+3}),1,0)")
    s.cell(row=6, column=COL0+j, value=f"={A}$B${JC_ROW0}*{cL}5").number_format = "0"
    s.cell(row=7, column=COL0+j, value=f"={A}{cL}${AU_ROW}").number_format = "#,##0"
    s.cell(row=8, column=COL0+j, value=f"={cL}7-{A}$B${JC_ROW0+1}").number_format = "#,##0"
    s.cell(row=9, column=COL0+j, value=f"={cL}6*{cL}8/1000").number_format = "#,##0"
    capex = JC["capex26"] if y == 2026 else (JC["capex27"] if y == 2027 else 0)
    cx = s.cell(row=10, column=COL0+j, value=(f"={A}$B${JC_ROW0+4}" if y == 2026 else (f"={A}$B${JC_ROW0+5}" if y == 2027 else 0)))
    cx.number_format = "#,##0"
    s.cell(row=11, column=COL0+j, value=f"={cL}9-{cL}10").number_format = "#,##0;(#,##0)"
    s.cell(row=12, column=COL0+j, value=f"=MAX(0,{cL}11*{A}$B${R_TAXUS})").number_format = "#,##0"
    s.cell(row=13, column=COL0+j, value=f"={cL}11-{cL}12").number_format = "#,##0;(#,##0)"
    s.cell(row=14, column=COL0+j, value=f"=1/(1+{A}$B${R_DISC})^({cL}4-0.5)").number_format = "0.000"
    s.cell(row=15, column=COL0+j, value=f"={cL}13*{cL}14").number_format = "#,##0;(#,##0)"
s["A17"] = "NPV ($mm)"; s["A17"].font = BOLD
s["B17"] = f"=SUM(D15:{get_column_letter(COL0+9)}15)"; s["B17"].font = BOLD; s["B17"].number_format = "#,##0"
s.column_dimensions["A"].width = 30

# ================= NAV summary =================
s = wb.create_sheet("NAV", 1)
s["A1"] = "Net Asset Value — summary"; s["A1"].font = Font(bold=True, size=13)
r = 3
s.cell(row=r, column=1, value="Asset").font = BOLD; s.cell(row=r, column=2, value="NPV ($mm)").font = BOLD
for c_ in ("A", "B"): s.cell(row=r, column={"A":1,"B":2}[c_]).fill = HDR
r += 1
for tab, (nm, *_rest) in zip(mine_tabs, MINES):
    s.cell(row=r, column=1, value=nm)
    c = s.cell(row=r, column=2, value=f"='{tab}'!B20"); c.font = GREEN; c.number_format = "#,##0"
    r += 1
s.cell(row=r, column=1, value="Jerritt Canyon (× include toggle)")
s.cell(row=r, column=2, value=f"=JerrittCanyon!B17*{A}B{R_JCON}").number_format = "#,##0"; s.cell(row=r, column=2).font = GREEN
JCR = r; r += 1
s.cell(row=r, column=1, value="Mining NAV").font = BOLD
s.cell(row=r, column=2, value=f"=SUM(B4:B{JCR})").number_format = "#,##0"; s.cell(row=r, column=2).font = BOLD
MNAV = r; r += 2
s.cell(row=r, column=1, value="Less: PV of corporate G&A (annuity)")
s.cell(row=r, column=2, value=f"=-{A}B{R_GA}*(1-(1+{A}B{R_DISC})^-{A}B{R_GAYRS})/{A}B{R_DISC}").number_format = "#,##0;(#,##0)"
GAR = r; r += 1
s.cell(row=r, column=1, value="Plus: net cash (Q1 2026)")
s.cell(row=r, column=2, value=f"={A}B{R_NETCASH}").number_format = "#,##0"; s.cell(row=r, column=2).font = GREEN
NCR = r; r += 1
s.cell(row=r, column=1, value="Corporate NAV").font = BOLD
s.cell(row=r, column=2, value=f"=B{MNAV}+B{GAR}+B{NCR}").number_format = "#,##0"; s.cell(row=r, column=2).font = BOLD
CNAV = r; r += 2
s.cell(row=r, column=1, value="Shares outstanding (M)")
s.cell(row=r, column=2, value=f"={A}B{R_SHARES}").number_format = "#,##0.0"; s.cell(row=r, column=2).font = GREEN
r += 1
s.cell(row=r, column=1, value="NAV per share (US$)").font = BOLD
s.cell(row=r, column=2, value=f"=B{CNAV}/B{r-1}").number_format = "#,##0.00"; s.cell(row=r, column=2).font = BOLD
NPS = r; r += 1
s.cell(row=r, column=1, value="Current price (US$)")
s.cell(row=r, column=2, value=f"={A}B{R_PRICE}").number_format = "#,##0.00"; s.cell(row=r, column=2).font = GREEN
r += 1
s.cell(row=r, column=1, value="P/NAV").font = BOLD
s.cell(row=r, column=2, value=f"=B{r-1}/B{NPS}").number_format = "0.00x"; s.cell(row=r, column=2).font = BOLD
r += 1
s.cell(row=r, column=1, value="Implied upside/(downside) to NAV")
s.cell(row=r, column=2, value=f"=B{NPS}/B{r-2}-1").number_format = "0.0%;(0.0%)"
s.column_dimensions["A"].width = 38; s.column_dimensions["B"].width = 14

wb.save("FirstMajestic_NAV.xlsx")
print("Saved FirstMajestic_NAV.xlsx")

# ================= Python mirror — sanity check =================
def mirror(deck, disc):
    ag_deck, au_deck = DECKS_AG[deck], DECKS_AU[deck]
    total = 0
    detail = {}
    for nm, ag, au, byp, aisc, rsv, ext in MINES:
        life = round(rsv / (ag + au/1000*67.5 + byp)) + ext
        npv = 0
        for t in range(1, 11):
            if t > life: break
            agp, aup = ag_deck[t-1], au_deck[t-1]
            ageq = ag + au/1000*aup/agp + byp
            rev = ag*agp + au/1000*aup + byp*agp
            cf = rev - aisc*ageq
            cf -= max(0, cf*CORP["tax_mx"])
            npv += cf / (1+disc)**(t-0.5)
        detail[nm] = npv; total += npv
    jc = 0
    for t in range(1, 11):
        y = 2025 + t
        cf = -JC["capex26"] if y == 2026 else (-JC["capex27"] if y == 2027 else 0)
        if JC["start"] <= y < JC["start"] + JC["years"]:
            cf += JC["au_koz"] * (au_deck[t-1] - JC["aisc_au"]) / 1000 * (1 - CORP["tax_us"])
        jc += cf / (1+disc)**(t-0.5)
    detail["Jerritt Canyon"] = jc; total += jc
    ga = -CORP["ga"] * (1-(1+disc)**-CORP["ga_years"]) / disc
    nav = total + ga + CORP["netcash"]
    return detail, nav, nav/CORP["shares"]

for deck in (1, 2, 3):
    d, nav, nps = mirror(deck, CORP["disc"])
    print(f"Deck {deck}: NAV ${nav:,.0f}M  |  NAV/share ${nps:.2f}  |  P/NAV {CORP['price']/nps:.2f}x")
d, nav, nps = mirror(2, CORP["disc"])
print("Mine NPVs (deck 2):", {k: round(v) for k, v in d.items()})
